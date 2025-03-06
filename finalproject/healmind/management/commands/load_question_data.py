import os.path
from django.contrib.auth.models import User
import pandas as pd
from django.core.management.base import BaseCommand
from healmind.models import *
from openpyxl import load_workbook
from django.conf import settings
from django.contrib.auth.models import Group



class Command(BaseCommand):
    help = 'Load data from an Excel file into the database'

    def handle(self, *args, **kwargs):
        # Load the Excel data
        file_path = os.path.join(settings.BASE_DIR,'healmind/fixtures/data3.2.xlsx')
        excel_data = pd.read_excel(file_path, sheet_name=None)

        # Load Excel workbook
        wb = load_workbook(filename=file_path)
        ws = wb.active  # สมมติว่า sheet แรกคือที่เราต้องการอ่าน

        member_group, created = Group.objects.get_or_create(name='member')
        doctor_group, created = Group.objects.get_or_create(name='doctor')

        ws = wb['member']
        for row in ws:
            values = [cell.value for cell in row]
            if values[0] != 'id':
                user = User.objects.create_user(values[1], values[2], str(values[8]), pk=values[0], first_name=values[3], last_name=values[4])
                member, created = Profile.objects.get_or_create(user=user, location=values[7],gender=values[5],age=values[6])

                user.groups.add(member_group)

        ws_doctor = wb['doctor']
        for row in ws_doctor:
            values = [cell.value for cell in row]
            if values[0] != 'id':

                password = str(values[6]) if values[6] else ''

                # Create user account for doctor
                user = User.objects.create_user(
                    username=values[4],
                    email=values[5],
                    password=password,
                    pk=values[0],
                    first_name=values[2],
                    last_name=values[3]
                )

                # Create doctor profile
                DoctorProfile.objects.create(
                    user=user,
                    title=values[1],
                    specialty=values[7],
                    bio=values[8],
                    session_rate=values[9],
                    work_location=values[10],
                    service_mode=values[11],
                    contact=values[12],
                    profile_image=values[13]
                )

                user.groups.add(doctor_group)

        ws_questionnaire = wb['questionnaire']
        for row in ws_questionnaire.iter_rows(min_row=2, values_only=True):
            questionnaire_id, name, description, is_system, created_by_id = row

            # ถ้ามี created_by_id ให้หา user object
            created_by = None
            if created_by_id:
                try:
                    created_by = User.objects.get(id=created_by_id)
                except User.DoesNotExist:
                    print(f"Warning: User with id {created_by_id} not found")

            questionnaire, created = Questionnaire.objects.get_or_create(
                id=questionnaire_id,
                defaults={
                    'questionnaire_name': name,
                    'description': description,
                    'is_system': bool(is_system),  # แปลงเป็น boolean
                    'created_by': created_by
                }
            )

        ws_questions = wb['question']
        for row in ws_questions.iter_rows(min_row=2, values_only=True):  # Skipping header row
            question_id, questionnaire_id, content = row

            # Get or create the questionnaire
            questionnaire = Questionnaire.objects.get(id=questionnaire_id)

            # Get or create the question
            question, created = Question.objects.get_or_create(
                id=question_id,
                defaults={'questionnaire': questionnaire, 'question_content': content}
            )

        # 4. Load "Choice" sheet
        ws_choice = wb['choice']
        for row in ws_choice.iter_rows(min_row=2, values_only=True):  # Skipping header row
            response_id, question_id, response_value, response_text = row
            question = Question.objects.get(id=question_id)

            # Get or create the choice
            choice, created = Choice.objects.get_or_create(
                id=response_id,
                defaults={'question': question, 'response_value': response_value, 'response_text': response_text}
            )

        # 5. Load "Results" sheet
        ws_results = wb['result']
        for row in ws_results.iter_rows(min_row=2, values_only=True):  # Skipping header row
            result_id, questionnaire_id, score_low, score_high, stress_level, result_description = row
            questionnaire = Questionnaire.objects.get(id=questionnaire_id)

            # Get or create the result
            result, created = Result.objects.get_or_create(
                id=result_id,
                defaults={'questionnaire': questionnaire, 'score_low': score_low, 'score_high': score_high,
                          'stress_level': stress_level, 'result_description': result_description}
            )

        ws_appointments = wb['appointments']
        for row in ws_appointments.iter_rows(min_row=2, values_only=True):
            doctor_id, member_id, appointment_date, time, service_mode, status, created_at, payment_status, stripe_payment_id = row

            try:
                # ดึง DoctorProfile โดยตรง
                doctor = DoctorProfile.objects.get(user_id=doctor_id)
                # ดึง Profile โดยตรง
                member = Profile.objects.get(user_id=member_id)

                Appointment.objects.create(
                    doctor=doctor,  # ใส่ DoctorProfile object แทน User
                    member=member,
                    appointment_date=appointment_date,
                    time=time,
                    service_mode=service_mode,
                    status=status,
                    created_at=created_at,
                    payment_status=payment_status,
                    stripe_payment_id=stripe_payment_id
                )
            except (DoctorProfile.DoesNotExist, Profile.DoesNotExist) as e:
                print(f"Error: {e}")
                continue

        self.stdout.write(self.style.SUCCESS('Data loaded successfully!'))

